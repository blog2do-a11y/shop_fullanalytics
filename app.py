#!/usr/bin/env python3
from flask import Flask, render_template, request, redirect, url_for, flash, send_from_directory, session, jsonify
import openpyxl, os, io, pandas as pd
from datetime import datetime
from werkzeug.utils import secure_filename

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'static', 'uploads')
EXCEL_FILE = os.path.join(BASE_DIR, 'orders.xlsx')
USERNAME = 'admin'
PASSWORD = 'Artur123!'
DELETE_CODE = os.environ.get('DELETE_CODE','Artur!')
ALLOWED_EXT = {'png','jpg','jpeg','gif'}

app = Flask(__name__)
app.secret_key = 'change_this_secret_key'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

HEADERS = [
    "Order ID","Full Name","First Name","Last Name","Weight (kg)","Height (cm)",
    "Address","Phone / Contact","Social Link","Order Link","Product Description","Comment",
    "Cost Price","Sale Price","Other Cost","Profit","Order DateTime","Order Month",
    "Shipping Method","Discount Type","Discount Value","Discount Notes","Payment Method",
    "Order Status","Image File"
]

def ensure_workbook():
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title='Orders'
        ws.append(HEADERS)
        wb.save(EXCEL_FILE)

def next_order_id():
    ensure_workbook()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    nums = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            nums.append(int(str(row[0]).split('-')[1]))
        except:
            pass
    n = max(nums)+1 if nums else 1
    wb.close()
    return f"ORD-{n:04d}"

def allowed_file(fn):
    return '.' in fn and fn.rsplit('.',1)[1].lower() in ALLOWED_EXT

def save_order(form, file):
    ensure_workbook()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    oid = next_order_id()
    now = datetime.now()
    ym = now.strftime('%Y-%m')
    folder = os.path.join(UPLOAD_FOLDER, ym)
    os.makedirs(folder, exist_ok=True)
    img_rel = ''
    if file and file.filename and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        filename = f"{now.strftime('%Y%m%d%H%M%S')}_{filename}"
        file.save(os.path.join(folder, filename))
        img_rel = f"{ym}/{filename}"
    def tofloat(x):
        try:
            return float(x)
        except:
            return 0.0
    cost = tofloat(form.get('cost_price'))
    sale = tofloat(form.get('sale_price'))
    other = tofloat(form.get('other_cost') or 0)
    profit = sale - cost - other
    order_dt = now.strftime('%Y-%m-%d %H:%M:%S')
    order_month = now.strftime('%Y-%m')
    row = [
        oid,
        f"{form.get('first_name','').strip()} {form.get('last_name','').strip()}".strip(),
        form.get('first_name',''),
        form.get('last_name',''),
        form.get('weight_kg',''),
        form.get('height_cm',''),
        form.get('address',''),
        form.get('phone',''),
        form.get('social_link',''),
        form.get('order_link',''),
        form.get('product_desc',''),
        form.get('comment',''),
        cost,
        sale,
        other,
        profit,
        order_dt,
        order_month,
        form.get('shipping_method',''),
        form.get('discount_type',''),
        form.get('discount_value',''),
        form.get('discount_notes',''),
        form.get('payment_method',''),
        form.get('order_status',''),
        img_rel
    ]
    ws.append(row)
    wb.save(EXCEL_FILE)
    wb.close()
    return oid

@app.route('/login', methods=['GET','POST'])
def login():
    if request.method == 'POST':
        u = request.form.get('username')
        p = request.form.get('password')
        if u == USERNAME and p == PASSWORD:
            session['logged_in'] = True
            return redirect(url_for('orders'))
        flash('Invalid credentials', 'error')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

def login_required(fn):
    from functools import wraps
    @wraps(fn)
    def wrap(*a, **k):
        if session.get('logged_in'):
            return fn(*a, **k)
        return redirect(url_for('login'))
    return wrap

@app.route('/')
@login_required
def root():
    return redirect(url_for('orders'))

@app.route('/add', methods=['GET','POST'])
@login_required
def add():
    if request.method == 'POST':
        oid = save_order(request.form, request.files.get('image'))
        flash(f'Order saved: {oid}', 'success')
        return redirect(url_for('orders'))
    return render_template('add.html')

@app.route('/orders')
@login_required
def orders():
    ensure_workbook()
    df = pd.read_excel(EXCEL_FILE, engine='openpyxl')
    df.fillna('', inplace=True)
    orders = df.to_dict(orient='records')
    return render_template('orders.html', orders=orders)

@app.route('/order/<order_id>')
@login_required
def order_detail(order_id):
    ensure_workbook()
    df = pd.read_excel(EXCEL_FILE, engine='openpyxl')
    df.fillna('', inplace=True)
    row = df[df['Order ID']==order_id]
    if row.empty:
        return jsonify({'ok': False}), 404
    rec = row.iloc[0].to_dict()
    rec2 = {k: (v.item() if hasattr(v,'item') else (v if not pd.isna(v) else '')) for k,v in rec.items()}
    return jsonify({'ok': True, 'order': rec2})

@app.route('/delete', methods=['POST'])
@login_required
def delete():
    payload = request.get_json() or {}
    order_id = payload.get('order_id')
    code = payload.get('code')
    if code != DELETE_CODE:
        return jsonify({'ok': False, 'error': 'Invalid code'}), 403
    ensure_workbook()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    found = False
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        if row[0].value == order_id:
            img = row[22].value if len(row)>22 else None
            if img:
                try:
                    img_path = os.path.join(UPLOAD_FOLDER, img)
                    if os.path.exists(img_path):
                        os.remove(img_path)
                except:
                    pass
            ws.delete_rows(idx,1)
            found = True
            break
    if found:
        wb.save(EXCEL_FILE)
    wb.close()
    return jsonify({'ok': found})

@app.route('/uploads/<path:filename>')
@login_required
def uploaded_file(filename):
    return send_from_directory(UPLOAD_FOLDER, filename)

def detect_platform(link):
    s = str(link).lower() if link else ''
    if 'instagram' in s: return 'Instagram'
    if 'facebook' in s or 'fb.me' in s: return 'Facebook'
    if 'tiktok' in s: return 'TikTok'
    if 'telegram' in s or 't.me' in s: return 'Telegram'
    if 'wa.me' in s or 'whatsapp' in s: return 'WhatsApp'
    if 'vk.com' in s: return 'VK'
    return 'Other' if s else 'Unknown'

@app.route('/stats')
@login_required
def stats():
    ensure_workbook()
    df = pd.read_excel(EXCEL_FILE, engine='openpyxl')
    if df.empty:
        monthly = []
        platform_counts = {}
        date_counts = []
    else:
        df.fillna('', inplace=True)
        df['Order Month'] = df['Order Month'].astype(str)
        months = sorted(df['Order Month'].unique())
        monthly = []
        for m in months:
            sub = df[df['Order Month']==m]
            count = int(sub.shape[0])
            revenue = float(sub['Sale Price'].astype(float).sum())
            cost = float(sub['Cost Price'].astype(float).sum())
            other = float(sub.get('Other Cost', 0).astype(float).sum()) if 'Other Cost' in sub else 0.0
            profit = revenue - cost - other
            monthly.append({'month': m, 'count': count, 'revenue': revenue, 'cost': cost, 'other': other, 'profit': profit})
        df['Platform'] = df['Social Link'].apply(detect_platform)
        platform_counts = df['Platform'].value_counts().to_dict()
        df['Order DateOnly'] = df['Order DateTime'].apply(lambda x: str(x).split(' ')[0] if x else '')
        date_counts_series = df['Order DateOnly'].value_counts().sort_index()
        date_counts = [{'date': d, 'count': int(c)} for d,c in date_counts_series.items()]
    return render_template('stats.html', monthly=monthly, platforms=platform_counts, date_counts=date_counts)

@app.route('/accounting')
@login_required
def accounting():
    ensure_workbook()
    df = pd.read_excel(EXCEL_FILE, engine='openpyxl')
    df.fillna('', inplace=True)
    month = request.args.get('month','')
    if month:
        df = df[df['Order Month']==month]
    for col in ['Cost Price','Sale Price','Other Cost','Profit']:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0.0)
    rows = df.to_dict(orient='records')
    months = sorted(pd.read_excel(EXCEL_FILE, engine='openpyxl')['Order Month'].dropna().unique())
    return render_template('accounting.html', rows=rows, months=months, selected_month=month)

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
